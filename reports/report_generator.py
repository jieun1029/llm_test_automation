"""
Excel 리포트 생성 모듈 (차트 포함)
"""
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import csv


def generate_excel_report(results, output_file="output/test_results.xlsx"):
    """
    Excel 보고서 생성 (2개 시트: 상세 결과 + 차트)
    
    Args:
        results (list): 테스트 결과 리스트
        output_file (str): 출력 파일 경로
    """
    wb = Workbook()
    
    # 기본 시트 삭제
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 시트 1: 상세 결과
    _create_detail_sheet(wb, results)
    
    # 시트 2: 차트
    _create_chart_sheet(wb, results)
    
    # 저장
    wb.save(output_file)


def _create_detail_sheet(wb, results):
    """시트1: 상세 결과 테이블"""
    ws = wb.create_sheet("상세 결과", 0)
    
    # 헤더 작성
    headers = [
        "질문ID", "라운드", "카테고리", "질문", 
        "정확성", "관련성", "할루시네이션", "안전성", "스타일", "기능적_요건",
        "총점", "만점", "PASS/FAIL", "타임스탬프"
    ]
    
    ws.append(headers)
    
    # 헤더 스타일링
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 데이터 작성
    for result in results:
        eval_data = result['evaluation']
        
        row_data = [
            result['test_id'],
            result.get('round', 1),
            result['category'],
            result['question'][:50] + "..." if len(result['question']) > 50 else result['question'],
            eval_data['scores']['정확성'],
            eval_data['scores']['관련성'],
            eval_data['scores']['할루시네이션'],
            eval_data['scores']['안전성'],
            eval_data['scores']['스타일'],
            eval_data['scores']['기능적_요건'],
            eval_data['total_score'],
            eval_data['max_score'],
            "PASS" if eval_data['pass'] else "FAIL",
            result['timestamp']
        ]
        
        ws.append(row_data)
    
    # PASS/FAIL 셀 색상 적용
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row_num in range(2, len(results) + 2):
        pass_fail_cell = ws.cell(row=row_num, column=13)
        if pass_fail_cell.value == "PASS":
            pass_fail_cell.fill = pass_fill
            pass_fail_cell.font = Font(color="006100", bold=True)
        else:
            pass_fail_cell.fill = fail_fill
            pass_fail_cell.font = Font(color="9C0006", bold=True)
        pass_fail_cell.alignment = Alignment(horizontal="center")
    
    # 점수 셀 가운데 정렬
    for row_num in range(2, len(results) + 2):
        for col_num in range(5, 12):  # 점수 컬럼들
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center")
    
    # 컬럼 너비 조정
    column_widths = [10, 8, 12, 40, 8, 8, 12, 8, 8, 12, 8, 8, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # 행 높이 조정
    ws.row_dimensions[1].height = 25
    for row_num in range(2, len(results) + 2):
        ws.row_dimensions[row_num].height = 20
    
    # 틀 고정 (헤더)
    ws.freeze_panes = "A2"


def _create_chart_sheet(wb, results):
    """시트2: 차트"""
    ws = wb.create_sheet("차트 & 통계", 1)
    
    # 통계 계산
    total_tests = len(results)
    passed = sum(1 for r in results if r['evaluation']['pass'])
    failed = total_tests - passed
    pass_rate = (passed / total_tests * 100) if total_tests > 0 else 0
    
    # 항목별 평균 점수
    criteria_names = ["정확성", "관련성", "할루시네이션", "안전성", "스타일", "기능적_요건"]
    avg_scores = {}
    
    for criterion in criteria_names:
        scores = [r['evaluation']['scores'][criterion] for r in results]
        avg_scores[criterion] = sum(scores) / len(scores) if scores else 0
    
    # === 1. 요약 통계 ===
    ws['A1'] = "📊 테스트 요약 통계"
    ws['A1'].font = Font(size=14, bold=True, color="4472C4")
    
    summary_data = [
        ["항목", "값"],
        ["총 테스트 수", total_tests],
        ["통과 (PASS)", passed],
        ["실패 (FAIL)", failed],
        ["통과율", f"{pass_rate:.1f}%"],
        ["평균 총점", f"{sum(r['evaluation']['total_score'] for r in results) / total_tests:.2f}/18점"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 3):
        ws.append(row_data)
        if row_idx == 3:  # 헤더
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    # === 2. PASS/FAIL 파이 차트 데이터 ===
    ws['D1'] = "PASS/FAIL 비율"
    ws['D1'].font = Font(size=12, bold=True)
    
    ws['D3'] = "상태"
    ws['E3'] = "개수"
    ws['D3'].font = Font(bold=True)
    ws['E3'].font = Font(bold=True)
    
    ws['D4'] = "PASS"
    ws['E4'] = passed
    ws['D5'] = "FAIL"
    ws['E5'] = failed
    
    # 파이 차트 생성
    pie = PieChart()
    labels = Reference(ws, min_col=4, min_row=4, max_row=5)
    data = Reference(ws, min_col=5, min_row=3, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "PASS/FAIL 비율"
    pie.height = 10
    pie.width = 15
    
    ws.add_chart(pie, "D7")
    
    # === 3. 항목별 평균 점수 막대 차트 데이터 ===
    ws['K1'] = "항목별 평균 점수"
    ws['K1'].font = Font(size=12, bold=True)
    
    ws['K3'] = "평가 항목"
    ws['L3'] = "평균 점수"
    ws['K3'].font = Font(bold=True)
    ws['L3'].font = Font(bold=True)
    
    row_start = 4
    for idx, (criterion, score) in enumerate(avg_scores.items(), row_start):
        ws.cell(row=idx, column=11).value = criterion
        ws.cell(row=idx, column=12).value = round(score, 2)
    
    # 막대 차트 생성
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "항목별 평균 점수 (만점 3점)"
    bar.y_axis.title = "점수"
    bar.x_axis.title = "평가 항목"
    
    data = Reference(ws, min_col=12, min_row=3, max_row=3 + len(avg_scores))
    cats = Reference(ws, min_col=11, min_row=4, max_row=3 + len(avg_scores))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 12
    bar.width = 20
    
    ws.add_chart(bar, "K10")
    
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 12


def save_detailed_answers_csv(results, output_file="output/detailed_answers.csv"):
    """답변 원문을 CSV로 저장"""
    
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # 헤더
        writer.writerow(['질문ID', '라운드', '카테고리', '질문', '답변 전문', '타임스탬프'])
        
        # 데이터
        for result in results:
            writer.writerow([
                result['test_id'],
                result.get('round', 1),
                result['category'],
                result['question'],
                result['answer'],
                result['timestamp']
            ])
